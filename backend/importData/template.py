"""
Génération du classeur Excel "modèle" que l'utilisateur télécharge, remplit
avec ses propres données, puis réimporte via `EquipementImportView`.

Le classeur contient un onglet par type de donnée, avec les colonnes déjà
nommées comme attendu par `importers.EquipementImporter`, et une ligne
d'exemple pré-remplie sous les en-têtes pour montrer le format attendu.
Cette ligne d'exemple est marquée par le suffixe "(EXEMPLE)" : l'import
l'ignore automatiquement (voir `importers._rows`), même si l'utilisateur
oublie de la supprimer avant d'importer son fichier.
"""

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from equipement.models import Equipement, StatutEquipement

EXEMPLE_MARQUEUR = "(EXEMPLE)"

# (nom_onglet, en_tetes, ligne_exemple)
SHEET_DEFINITIONS = [
    (
        "Lieux",
        ["Nom *", "Type de lieu", "Lieu parent", "Lien du plan"],
        [f"Atelier Mecanique {EXEMPLE_MARQUEUR}", "Atelier", "", ""],
    ),
    (
        "Fabricants",
        [
            "Nom *",
            "Email",
            "Telephone",
            "Service apres-vente (Oui/Non)",
            "N adresse",
            "Rue",
            "Ville",
            "Code postal",
            "Pays",
            "Complement adresse",
        ],
        [
            f"Siemens AG {EXEMPLE_MARQUEUR}", "contact@siemens.fr", "0100000000", "Oui",
            "12", "Rue de l'Industrie", "Anglet", "64600", "France", "",
        ],
    ),
    (
        "Fournisseurs",
        [
            "Nom *",
            "Email",
            "Telephone",
            "Service apres-vente (Oui/Non)",
            "N adresse",
            "Rue",
            "Ville",
            "Code postal",
            "Pays",
            "Complement adresse",
        ],
        [
            f"Distrelec {EXEMPLE_MARQUEUR}", "contact@distrelec.fr", "0200000000", "Non",
            "5", "Avenue des Fournisseurs", "Bayonne", "64100", "France", "",
        ],
    ),
    (
        "Familles",
        ["Nom *", "Famille parente"],
        [f"Moteurs electriques {EXEMPLE_MARQUEUR}", f"Equipements electriques {EXEMPLE_MARQUEUR}"],
    ),
    (
        "Modeles",
        ["Nom *", "Fabricant *"],
        [f"S7-1200 {EXEMPLE_MARQUEUR}", f"Siemens AG {EXEMPLE_MARQUEUR}"],
    ),
    (
        "Equipements",
        [
            "Code GMAO *",
            "Designation *",
            "Type",
            "Statut *",
            "Lieu *",
            "Fabricant",
            "Fournisseur",
            "Famille",
            "Modele",
            "Date de mise en service (JJ/MM/AAAA)",
            "Prix d'achat",
            "N de serie",
        ],
        [
            f"EQ-001 {EXEMPLE_MARQUEUR}",
            f"Perceuse a colonne {EXEMPLE_MARQUEUR}",
            "MECANIQUE",
            "EN_FONCTIONNEMENT",
            f"Atelier Mecanique {EXEMPLE_MARQUEUR}",
            f"Siemens AG {EXEMPLE_MARQUEUR}",
            f"Distrelec {EXEMPLE_MARQUEUR}",
            f"Equipements electriques {EXEMPLE_MARQUEUR}",
            f"S7-1200 {EXEMPLE_MARQUEUR}",
            "15/03/2023",
            "1250.50",
            "FR-000123",
        ],
    ),
]


def _build_instructions_sheet(workbook):
    sheet = workbook.create_sheet(title="Instructions", index=0)
    bold = Font(bold=True)

    sheet.cell(row=1, column=1, value="Comment remplir ce fichier").font = Font(bold=True, size=14)

    regles = [
        "1. Remplissez un onglet par type de donnee (Lieux, Fabricants, Fournisseurs, Familles, Modeles, Equipements).",
        "2. Les colonnes marquees d'un asterisque (*) sont obligatoires, les autres sont facultatives.",
        "3. Si un lieu, un fabricant, un fournisseur ou une famille indique dans l'onglet Equipements n'existe pas "
        "encore, il sera cree automatiquement lors de l'import.",
        "4. N'ajoutez pas, ne renommez pas et ne supprimez pas les onglets ou les colonnes.",
        f"5. Valeurs possibles pour la colonne Type (onglet Equipements) : "
        f"{', '.join(code for code, _ in Equipement.TYPE_CHOICES)}.",
        f"6. Valeurs possibles pour la colonne Statut (onglet Equipements) : "
        f"{', '.join(code for code, _ in StatutEquipement.STATUTS_CHOICES)}.",
        "7. Le Code GMAO identifie chaque equipement de maniere unique : reimporter un fichier avec les memes "
        "codes GMAO ne creera pas de doublons (les lignes deja presentes sont ignorees).",
        "8. Les dates doivent etre au format JJ/MM/AAAA.",
        "9. Chaque onglet contient une ligne d'exemple (EXEMPLE) montrant le format attendu : "
        "elle est ignoree automatiquement lors de l'import, meme si vous oubliez de la supprimer.",
        "10. Pour Service apres-vente, utilisez Oui ou Non (Non par defaut si laisse vide).",
        "11. L'adresse (N adresse, Rue, Ville, Code postal, Pays, Complement) est facultative pour "
        "un fabricant ou un fournisseur : elle n'est enregistree que si au moins un de ces champs est rempli.",
        "12. La photo d'un equipement et les consommables associes ne sont PAS geres par cet import "
        "(l'un est une image, l'autre necessite des quantites) : ajoutez-les ensuite manuellement sur la fiche "
        "de l'equipement une fois cree.",
        "13. Une fois rempli, importez ce fichier depuis la page Equipements, bouton 'Importer depuis Excel'.",
    ]
    row = 3
    for ligne in regles:
        sheet.cell(row=row, column=1, value=ligne)
        row += 1

    # Tableau de référence : une ligne par colonne attendue, avec son exemple
    row += 1
    headers = ["Onglet", "Colonne", "Obligatoire", "Exemple"]
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=col_idx, value=header)
        cell.font = bold
    row += 1

    for sheet_name, sheet_headers, exemple_row in SHEET_DEFINITIONS:
        for header, exemple_valeur in zip(sheet_headers, exemple_row):
            obligatoire = "Oui" if header.strip().endswith("*") else "Non"
            colonne = header.replace("*", "").strip()
            valeur_affichee = str(exemple_valeur).replace(EXEMPLE_MARQUEUR, "").strip()
            sheet.cell(row=row, column=1, value=sheet_name)
            sheet.cell(row=row, column=2, value=colonne)
            sheet.cell(row=row, column=3, value=obligatoire)
            sheet.cell(row=row, column=4, value=valeur_affichee)
            row += 1

    for col_letter, width in (("A", 22), ("B", 40), ("C", 14), ("D", 30)):
        sheet.column_dimensions[col_letter].width = width


def generate_equipement_template():
    """Retourne une HttpResponse contenant le classeur .xlsx modèle."""
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # on retire l'onglet vide créé par défaut

    header_font = Font(bold=True)
    exemple_font = Font(italic=True, color="808080")
    exemple_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for sheet_name, headers, exemple_row in SHEET_DEFINITIONS:
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(headers)
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = header_font
            sheet.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 22)

        sheet.append(exemple_row)
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=2, column=col_idx)
            cell.font = exemple_font
            cell.fill = exemple_fill

    _build_instructions_sheet(workbook)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="modele_import_equipements.xlsx"'
    workbook.save(response)
    return response
